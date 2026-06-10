"""Admin marketing campaign routes — bulk email with automatic domain warm-up."""

from __future__ import annotations

import io
import uuid
from datetime import date, datetime, timezone

import openpyxl
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.api.deps.auth import get_current_admin
from app.api.deps.db import get_db
from app.models.marketing_campaign import MarketingCampaign, MarketingCampaignContact
from app.models.mailing_list import Recipient, RecipientList
from app.models.user import User
from app.services.brevo_service import (
    _WARMUP_SCHEDULE,
    get_all_clickers,
    get_all_openers,
    get_brevo_api_key,
    get_email_campaigns,
    get_warmup_daily_limit,
)

router = APIRouter(prefix="/admin/marketing", tags=["admin-marketing"])


# ── Schemas ────────────────────────────────────────────────────────────────────

class CampaignCreate(BaseModel):
    name: str
    subject: str
    html_body: str
    from_name: str = "JobAI24"
    from_email: str = "marketing@jobai24.com"


class CampaignResponse(BaseModel):
    id: str
    name: str
    subject: str
    from_name: str
    from_email: str
    status: str
    warmup_start_date: date | None
    current_daily_limit: int
    total_contacts: int
    total_sent: int
    total_failed: int
    last_sent_at: datetime | None
    completed_at: datetime | None
    error_message: str | None
    created_at: datetime
    updated_at: datetime
    progress_pct: float

    class Config:
        from_attributes = True


def _to_response(c: MarketingCampaign) -> CampaignResponse:
    pct = round(c.total_sent / c.total_contacts * 100, 1) if c.total_contacts else 0.0
    return CampaignResponse(
        id=c.id,
        name=c.name,
        subject=c.subject,
        from_name=c.from_name,
        from_email=c.from_email,
        status=c.status,
        warmup_start_date=c.warmup_start_date,
        current_daily_limit=c.current_daily_limit,
        total_contacts=c.total_contacts,
        total_sent=c.total_sent,
        total_failed=c.total_failed,
        last_sent_at=c.last_sent_at,
        completed_at=c.completed_at,
        error_message=c.error_message,
        created_at=c.created_at,
        updated_at=c.updated_at,
        progress_pct=pct,
    )


# ── Endpoints ──────────────────────────────────────────────────────────────────

@router.post("/campaigns", response_model=CampaignResponse, status_code=status.HTTP_201_CREATED)
def create_campaign(
    body: CampaignCreate,
    admin: User = Depends(get_current_admin),
    db: Session = Depends(get_db),
) -> CampaignResponse:
    campaign = MarketingCampaign(
        id=str(uuid.uuid4()),
        name=body.name,
        subject=body.subject,
        html_body=body.html_body,
        from_name=body.from_name,
        from_email=body.from_email,
        status="draft",
    )
    db.add(campaign)
    db.commit()
    db.refresh(campaign)
    return _to_response(campaign)


@router.get("/campaigns", response_model=list[CampaignResponse])
def list_campaigns(
    admin: User = Depends(get_current_admin),
    db: Session = Depends(get_db),
) -> list[CampaignResponse]:
    campaigns = db.scalars(
        select(MarketingCampaign).order_by(MarketingCampaign.created_at.desc())
    ).all()
    return [_to_response(c) for c in campaigns]


@router.get("/campaigns/{campaign_id}", response_model=CampaignResponse)
def get_campaign(
    campaign_id: str,
    admin: User = Depends(get_current_admin),
    db: Session = Depends(get_db),
) -> CampaignResponse:
    campaign = db.get(MarketingCampaign, campaign_id)
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found.")
    return _to_response(campaign)


@router.post("/campaigns/{campaign_id}/import", response_model=dict)
async def import_contacts(
    campaign_id: str,
    file: UploadFile = File(...),
    admin: User = Depends(get_current_admin),
    db: Session = Depends(get_db),
) -> dict:
    campaign = db.get(MarketingCampaign, campaign_id)
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found.")
    if campaign.status not in ("draft",):
        raise HTTPException(status_code=400, detail="Can only import contacts to a draft campaign.")

    filename = (file.filename or "").lower()
    if not (filename.endswith(".xlsx") or filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="Only .xlsx or .xls files are accepted.")

    raw = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb.active
    except Exception:
        raise HTTPException(status_code=400, detail="Could not read Excel file.")

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"added": 0, "skipped": 0}

    header = [str(c).strip().lower() if c else "" for c in rows[0]]
    col: dict[str, int] = {}
    for name in ("email", "full_name"):
        if name in header:
            col[name] = header.index(name)

    if "email" not in col:
        raise HTTPException(status_code=400, detail="Excel must have an 'email' column.")

    added = skipped = 0
    for row in rows[1:]:
        email_val = row[col["email"]] if col["email"] < len(row) else None
        if not email_val:
            skipped += 1
            continue
        email_str = str(email_val).strip().lower()
        if "@" not in email_str:
            skipped += 1
            continue

        full_name_val = None
        if "full_name" in col and col["full_name"] < len(row):
            v = row[col["full_name"]]
            full_name_val = str(v).strip() if v else None

        db.add(MarketingCampaignContact(
            id=str(uuid.uuid4()),
            campaign_id=campaign_id,
            email=email_str,
            full_name=full_name_val,
        ))
        added += 1

    campaign.total_contacts += added
    db.commit()
    return {"added": added, "skipped": skipped}


@router.get("/campaigns/template/download")
def download_template(admin: User = Depends(get_current_admin)) -> StreamingResponse:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contacts"
    ws.append(["email", "full_name"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=marketing_contacts_template.xlsx"},
    )


@router.patch("/campaigns/{campaign_id}/activate", response_model=CampaignResponse)
def activate_campaign(
    campaign_id: str,
    admin: User = Depends(get_current_admin),
    db: Session = Depends(get_db),
) -> CampaignResponse:
    campaign = db.get(MarketingCampaign, campaign_id)
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found.")
    if campaign.total_contacts == 0:
        raise HTTPException(status_code=400, detail="Import contacts before activating the campaign.")
    if campaign.status not in ("draft", "paused"):
        raise HTTPException(status_code=400, detail=f"Cannot activate a campaign with status '{campaign.status}'.")
    campaign.status = "active"
    db.commit()
    db.refresh(campaign)
    return _to_response(campaign)


@router.patch("/campaigns/{campaign_id}/pause", response_model=CampaignResponse)
def pause_campaign(
    campaign_id: str,
    admin: User = Depends(get_current_admin),
    db: Session = Depends(get_db),
) -> CampaignResponse:
    campaign = db.get(MarketingCampaign, campaign_id)
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found.")
    if campaign.status != "active":
        raise HTTPException(status_code=400, detail="Only active campaigns can be paused.")
    campaign.status = "paused"
    db.commit()
    db.refresh(campaign)
    return _to_response(campaign)


@router.patch("/campaigns/{campaign_id}/resume", response_model=CampaignResponse)
def resume_campaign(
    campaign_id: str,
    admin: User = Depends(get_current_admin),
    db: Session = Depends(get_db),
) -> CampaignResponse:
    campaign = db.get(MarketingCampaign, campaign_id)
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found.")
    if campaign.status != "paused":
        raise HTTPException(status_code=400, detail="Only paused campaigns can be resumed.")
    campaign.status = "active"
    db.commit()
    db.refresh(campaign)
    return _to_response(campaign)


@router.get("/warmup-schedule")
def get_warmup_schedule(_: User = Depends(get_current_admin)) -> list[dict]:
    """Return the automatic warm-up schedule for UI display."""
    return [
        {"min_day": mn, "max_day": mx if mx < 9999 else None, "daily_limit": lim}
        for mn, mx, lim in _WARMUP_SCHEDULE
    ]


# ── Brevo Analytics Endpoints ──────────────────────────────────────────────────

@router.get("/brevo/campaigns")
async def list_brevo_campaigns(
    admin: User = Depends(get_current_admin),
) -> list[dict]:
    """Fetch all sent campaigns from Brevo with their open/click stats."""
    if not get_brevo_api_key():
        raise HTTPException(status_code=503, detail="BREVO_API_KEY not configured.")
    try:
        # Fetch up to 100 campaigns (paginate if needed)
        data = await get_email_campaigns(limit=50, offset=0)
        campaigns = data.get("campaigns", [])
        total = data.get("count", 0)
        if total > 50:
            more = await get_email_campaigns(limit=50, offset=50)
            campaigns.extend(more.get("campaigns", []))

        result = []
        for c in campaigns:
            campaign_stats = (c.get("statistics") or {}).get("campaignStats") or []
            delivered      = sum(s.get("delivered", 0)      for s in campaign_stats)
            unique_opens   = sum(s.get("uniqueViews", 0)    for s in campaign_stats)
            unique_clicks  = sum(s.get("uniqueClicks", 0)   for s in campaign_stats)
            unsubscriptions = sum(s.get("unsubscriptions", 0) for s in campaign_stats)
            result.append({
                "id": c.get("id"),
                "name": c.get("name", ""),
                "subject": c.get("subject", ""),
                "sent_date": c.get("sentDate"),
                "delivered": delivered,
                "unique_opens": unique_opens,
                "unique_clicks": unique_clicks,
                "unsubscriptions": unsubscriptions,
                "open_rate": round(unique_opens / delivered * 100, 1) if delivered else 0,
            })
        return result
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"Brevo API error: {exc}")


class SaveOpenersBody(BaseModel):
    list_name: str
    include_clickers: bool = False


@router.get("/brevo/campaigns/{campaign_id}/openers")
async def get_brevo_openers(
    campaign_id: int,
    admin: User = Depends(get_current_admin),
) -> dict:
    """Preview openers count for a Brevo campaign."""
    if not get_brevo_api_key():
        raise HTTPException(status_code=503, detail="BREVO_API_KEY not configured.")
    try:
        openers = await get_all_openers(campaign_id)
        return {"campaign_id": campaign_id, "count": len(openers)}
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"Brevo API error: {exc}")


@router.post("/brevo/campaigns/{campaign_id}/save-openers")
async def save_openers_to_list(
    campaign_id: int,
    body: SaveOpenersBody,
    admin: User = Depends(get_current_admin),
    db: Session = Depends(get_db),
) -> dict:
    """Extract openers (and optionally clickers) from a Brevo campaign and save to a new RecipientList."""
    if not get_brevo_api_key():
        raise HTTPException(status_code=503, detail="BREVO_API_KEY not configured.")
    if not body.list_name.strip():
        raise HTTPException(status_code=400, detail="list_name is required.")

    try:
        openers = await get_all_openers(campaign_id)
        contacts: dict[str, str | None] = {
            item["email"]: item.get("name") or None
            for item in openers
            if item.get("email")
        }

        if body.include_clickers:
            clickers = await get_all_clickers(campaign_id)
            for item in clickers:
                email = item.get("email")
                if email and email not in contacts:
                    contacts[email] = item.get("name") or None

    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"Brevo API error: {exc}")

    if not contacts:
        raise HTTPException(status_code=404, detail="No openers found for this campaign.")

    # Create a new RecipientList owned by the admin
    new_list = RecipientList(
        id=str(uuid.uuid4()),
        user_id=admin.id,
        name=body.list_name.strip(),
        description=f"Extracted from Brevo campaign #{campaign_id}",
        source="import",
        total_count=0,
    )
    db.add(new_list)
    db.flush()

    added = 0
    for email, name in contacts.items():
        db.add(Recipient(
            id=str(uuid.uuid4()),
            list_id=new_list.id,
            user_id=admin.id,
            email=email.lower().strip(),
            full_name=name,
        ))
        added += 1

    new_list.total_count = added
    db.commit()

    return {
        "list_id": new_list.id,
        "list_name": new_list.name,
        "added": added,
    }
