"""Parse GOSI (Social Insurance) Excel reports and compute Saudization summary stats."""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

# Common Arabic/English column name aliases for GOSI exports
_NAME_ALIASES = {"الاسم", "اسم الموظف", "اسم المشترك", "المشترك", "name", "employee name", "full name", "subscriber name"}
_ID_ALIASES = {"رقم الهوية", "رقم الإقامة", "هوية", "national id", "iqama", "id number", "رقم الوثيقة"}
_NATIONALITY_ALIASES = {"الجنسية", "nationality", "جنسية"}
_JOB_ALIASES = {"المهنة", "المسمى الوظيفي", "job title", "profession", "occupation", "وظيفة", "مهنة"}
_GENDER_ALIASES = {"الجنس", "gender", "sex", "جنس"}
_HIRE_DATE_ALIASES = {"تاريخ التعيين", "تاريخ الالتحاق", "تاريخ الإلتحاق", "تاريخ الالتحاق بالعمل",
                       "hire date", "start date", "joining date", "join date"}
_BIRTH_DATE_ALIASES = {"تاريخ الميلاد", "الميلاد", "birth date", "date of birth", "dob"}
_BASIC_SALARY_ALIASES = {"الأجر الأساسي", "الراتب الأساسي", "أجر أساسي", "basic salary", "basic wage", "basic"}
_HOUSING_ALIASES = {"السكن", "بدل سكن", "housing", "housing allowance"}
_VARIABLES_ALIASES = {"المتغيرات", "متغيرات", "variables", "variable pay"}
_ALLOWANCES_ALIASES = {"البدلات", "بدلات", "allowances", "other allowances"}
_TOTAL_WAGE_ALIASES = {"الأجر الإجمالي", "الراتب الإجمالي", "إجمالي الراتب", "total wage", "total salary", "gross salary"}
_SPECIAL_WAGE_ALIASES = {"الأجر الخاص", "راتب خاص", "special wage", "special salary"}

SAUDI_NATIONALITY_VALUES = {"سعودي", "سعودية", "saudi", "saudi arabia", "sa", "سعودي/ة"}


def _normalize(val: str) -> str:
    return str(val).strip().lower()


def _match_column(col_name: str, aliases: set[str]) -> bool:
    n = _normalize(col_name)
    return any(alias in n or n in alias for alias in aliases)


def _find_column(headers: list[str], aliases: set[str]) -> int | None:
    for i, h in enumerate(headers):
        if _match_column(str(h), aliases):
            return i
    return None


def validate_gosi_excel_bytes(file_bytes: bytes) -> list[str]:
    """
    Quickly validate that file_bytes looks like a GOSI Excel with required columns.
    Returns a list of Arabic error strings (empty list = valid).
    """
    try:
        import openpyxl
        from io import BytesIO
        wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(max_row=12, values_only=True))
        wb.close()
    except Exception as e:
        return [f"لا يمكن قراءة الملف: {e}"]

    if not rows:
        return ["الملف فارغ."]

    # Look for a header row in first 12 rows
    for row in rows:
        row_strs = [str(c) if c is not None else "" for c in row]
        has_name = any(_match_column(h, _NAME_ALIASES) for h in row_strs)
        has_nat  = any(_match_column(h, _NATIONALITY_ALIASES) for h in row_strs)
        has_job  = any(_match_column(h, _JOB_ALIASES) for h in row_strs)
        if has_name or has_nat or has_job:
            errors: list[str] = []
            if not has_name:
                errors.append("عمود اسم الموظف/المشترك غير موجود — المتوقع: «اسم المشترك» أو «الاسم» أو «Employee Name»")
            if not has_nat:
                errors.append("عمود الجنسية غير موجود — المتوقع: «الجنسية» أو «Nationality»")
            if not has_job:
                errors.append("عمود المهنة غير موجود — المتوقع: «المهنة» أو «المسمى الوظيفي» أو «Job Title»")
            return errors

    return [
        "لم يُعثر على صف رؤوس الأعمدة في أول 12 صفاً. "
        "تأكد أن الملف يحتوي على بيانات موظفين من التأمينات الاجتماعية (GOSI) "
        "وأن الأعمدة المطلوبة هي: اسم المشترك، الجنسية، المهنة."
    ]


def parse_gosi_excel(file_path: Path) -> dict[str, Any]:
    """
    Parse a GOSI Excel file and return employees list + summary stats.

    Returns:
        {
            "employees": [...],
            "summary": {
                "total": int,
                "saudi_count": int,
                "non_saudi_count": int,
                "saudization_pct": float,
                "by_profession": {title: {"total": int, "saudi": int, "pct": float}}
            }
        }
    """
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl is required for Excel parsing. Add it to requirements.txt.")

    wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"employees": [], "summary": _empty_summary()}

    # Find header row (first row with recognizable column)
    header_row_idx = 0
    headers: list[str] = []
    for idx, row in enumerate(rows[:10]):
        row_strs = [str(c) if c is not None else "" for c in row]
        if any(_match_column(h, _NATIONALITY_ALIASES | _JOB_ALIASES) for h in row_strs):
            header_row_idx = idx
            headers = row_strs
            break
    else:
        headers = [str(c) if c is not None else f"col_{i}" for i, c in enumerate(rows[0])]

    name_idx         = _find_column(headers, _NAME_ALIASES)
    id_idx           = _find_column(headers, _ID_ALIASES)
    nat_idx          = _find_column(headers, _NATIONALITY_ALIASES)
    job_idx          = _find_column(headers, _JOB_ALIASES)
    gender_idx       = _find_column(headers, _GENDER_ALIASES)
    hire_idx         = _find_column(headers, _HIRE_DATE_ALIASES)
    birth_idx        = _find_column(headers, _BIRTH_DATE_ALIASES)
    basic_sal_idx    = _find_column(headers, _BASIC_SALARY_ALIASES)
    housing_idx      = _find_column(headers, _HOUSING_ALIASES)
    variables_idx    = _find_column(headers, _VARIABLES_ALIASES)
    allowances_idx   = _find_column(headers, _ALLOWANCES_ALIASES)
    total_wage_idx   = _find_column(headers, _TOTAL_WAGE_ALIASES)
    special_wage_idx = _find_column(headers, _SPECIAL_WAGE_ALIASES)

    import datetime as _dt

    def _fmt_date(v: Any) -> str | None:
        if v is None:
            return None
        if isinstance(v, (_dt.date, _dt.datetime)):
            return v.strftime("%Y-%m-%d")
        s = str(v).strip()
        # strip time portion "2025-01-01 00:00:00"
        if " " in s and s.count("-") >= 2:
            return s.split(" ")[0]
        return s or None

    def _fmt_num(v: Any) -> float | None:
        if v is None:
            return None
        try:
            return float(str(v).replace(",", "").strip())
        except (ValueError, TypeError):
            return None

    employees: list[dict] = []
    for row in rows[header_row_idx + 1:]:
        if not any(row):
            continue

        def get(idx: int | None) -> str | None:
            if idx is None or idx >= len(row):
                return None
            v = row[idx]
            if v is None:
                return None
            if isinstance(v, (_dt.date, _dt.datetime)):
                return v.strftime("%Y-%m-%d")
            return str(v).strip() or None

        def getr(idx: int | None) -> Any:
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        nationality = get(nat_idx) or ""
        employees.append({
            "name":         get(name_idx),
            "national_id":  get(id_idx),
            "nationality":  nationality,
            "job_title":    get(job_idx),
            "gender":       get(gender_idx),
            "hire_date":    _fmt_date(getr(hire_idx)),
            "birth_date":   _fmt_date(getr(birth_idx)),
            "basic_salary": _fmt_num(getr(basic_sal_idx)),
            "housing":      _fmt_num(getr(housing_idx)),
            "variables":    _fmt_num(getr(variables_idx)),
            "allowances":   _fmt_num(getr(allowances_idx)),
            "total_wage":   _fmt_num(getr(total_wage_idx)),
            "special_wage": _fmt_num(getr(special_wage_idx)),
            "is_saudi": _normalize(nationality) in {_normalize(v) for v in SAUDI_NATIONALITY_VALUES},
        })

    wb.close()
    summary = _compute_summary(employees)
    return {"employees": employees, "summary": summary}


def _compute_summary(employees: list[dict]) -> dict[str, Any]:
    total = len(employees)
    saudi_count = sum(1 for e in employees if e.get("is_saudi"))
    non_saudi_count = total - saudi_count
    saudization_pct = round((saudi_count / total * 100), 2) if total > 0 else 0.0

    by_profession: dict[str, dict] = {}
    for emp in employees:
        title = (emp.get("job_title") or "غير محدد").strip() or "غير محدد"
        if title not in by_profession:
            by_profession[title] = {"total": 0, "saudi": 0, "pct": 0.0}
        by_profession[title]["total"] += 1
        if emp.get("is_saudi"):
            by_profession[title]["saudi"] += 1

    for title, stats in by_profession.items():
        t = stats["total"]
        stats["pct"] = round((stats["saudi"] / t * 100), 2) if t > 0 else 0.0

    return {
        "total": total,
        "saudi_count": saudi_count,
        "non_saudi_count": non_saudi_count,
        "saudization_pct": saudization_pct,
        "by_profession": by_profession,
    }


def _empty_summary() -> dict[str, Any]:
    return {"total": 0, "saudi_count": 0, "non_saudi_count": 0, "saudization_pct": 0.0, "by_profession": {}}
